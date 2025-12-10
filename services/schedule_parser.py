from pathlib import Path
from typing import List, Optional, Dict
from datetime import date, time
import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from models.timeslot import TimeSlot
from config.constants import SlotStatus
from utils.date_utils import parse_date, parse_time_range
from utils.logger import get_logger

logger = get_logger(__name__)


class ScheduleParser:
    def __init__(self, file_path: str):
        self.file_path = Path(file_path)
        self.workbook: Optional[Workbook] = None
        self.worksheet: Optional[Worksheet] = None
    
    def load(self) -> bool:
        try:
            if not self.file_path.exists():
                logger.warning(f"Файл расписания не найден: {self.file_path}")
                self._create_new_schedule()
                return True
            
            self.workbook = openpyxl.load_workbook(self.file_path)
            
            if self.workbook.sheetnames:
                self.worksheet = self.workbook.active
            else:
                self.worksheet = self.workbook.create_sheet("Расписание")
            
            logger.info(f"Расписание загружено из {self.file_path}")
            return True
            
        except Exception as e:
            logger.error(f"Ошибка при загрузке расписания: {e}", exc_info=True)
            return False
    
    def save(self) -> bool:
        try:
            if self.workbook is None:
                logger.error("Нет загруженной книги для сохранения")
                return False
            
            self.file_path.parent.mkdir(parents=True, exist_ok=True)
            
            self.workbook.save(self.file_path)
            logger.info(f"Расписание сохранено в {self.file_path}")
            return True
            
        except Exception as e:
            logger.error(f"Ошибка при сохранении расписания: {e}", exc_info=True)
            return False
    
    def _create_new_schedule(self) -> None:
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "Расписание"
        
        headers = ["Дата", "Время", "Статус", "Эксперт ID", "Имя эксперта"]
        self.worksheet.append(headers)
        
        logger.info("Создан новый файл расписания")
    
    def read_all_slots(self) -> List[TimeSlot]:
        if self.worksheet is None:
            logger.error("Таблица не загружена")
            return []
        
        slots = []
        
        for row in self.worksheet.iter_rows(min_row=2, values_only=True):
            try:
                slot = self._parse_row_to_slot(row)
                if slot:
                    slots.append(slot)
            except Exception as e:
                logger.warning(f"Ошибка при парсинге строки {row}: {e}")
                continue
        
        logger.info(f"Прочитано {len(slots)} слотов из расписания")
        return slots
    
    def _parse_row_to_slot(self, row: tuple) -> Optional[TimeSlot]:
        if not row or len(row) < 3:
            return None
        
        date_str = str(row[0]) if row[0] else None
        time_range_str = str(row[1]) if row[1] else None
        status_str = str(row[2]) if row[2] else "free"
        expert_id = int(row[3]) if len(row) > 3 and row[3] else None
        expert_name = str(row[4]) if len(row) > 4 and row[4] else None
        
        if not date_str or not time_range_str:
            return None
        
        slot_date = parse_date(date_str)
        if not slot_date:
            return None
        
        time_range = parse_time_range(time_range_str)
        if not time_range:
            return None
        
        start_time, end_time = time_range
        
        try:
            status = SlotStatus(status_str.lower())
        except ValueError:
            status = SlotStatus.FREE
        
        return TimeSlot(
            slot_date=slot_date,
            start_time=start_time,
            end_time=end_time,
            status=status,
            expert_id=expert_id,
            expert_name=expert_name
        )
    
    def write_slots(self, slots: List[TimeSlot], clear_existing: bool = False) -> bool:
        try:
            if self.worksheet is None:
                self._create_new_schedule()
            
            if clear_existing:
                self.worksheet.delete_rows(2, self.worksheet.max_row)
            
            for slot in slots:
                row_data = self._slot_to_row(slot)
                self.worksheet.append(row_data)
            
            logger.info(f"Записано {len(slots)} слотов в расписание")
            return True
            
        except Exception as e:
            logger.error(f"Ошибка при записи слотов: {e}", exc_info=True)
            return False
    
    def _slot_to_row(self, slot: TimeSlot) -> List:
        return [
            slot.format_date(),
            slot.format_time_range(),
            slot.status.value,
            slot.expert_id if slot.expert_id else "",
            slot.expert_name if slot.expert_name else ""
        ]
    
    def update_slot_status(
        self,
        slot_date: date,
        time_range: str,
        new_status: SlotStatus,
        expert_id: Optional[int] = None,
        expert_name: Optional[str] = None
    ) -> bool:
        try:
            if self.worksheet is None:
                logger.error("Таблица не загружена")
                return False
            
            date_str = slot_date.strftime("%d.%m.%Y")
            
            for row_idx, row in enumerate(self.worksheet.iter_rows(min_row=2), start=2):
                cell_date = str(row[0].value) if row[0].value else ""
                cell_time = str(row[1].value) if row[1].value else ""
                
                if cell_date == date_str and cell_time == time_range:
                    self.worksheet.cell(row=row_idx, column=3, value=new_status.value)
                    
                    if expert_id is not None:
                        self.worksheet.cell(row=row_idx, column=4, value=expert_id)
                    if expert_name is not None:
                        self.worksheet.cell(row=row_idx, column=5, value=expert_name)
                    
                    logger.info(f"Обновлён слот {date_str} {time_range} → {new_status.value}")
                    return True
            
            logger.warning(f"Слот не найден: {date_str} {time_range}")
            return False
            
        except Exception as e:
            logger.error(f"Ошибка при обновлении слота: {e}", exc_info=True)
            return False
    
    def get_free_slots(self) -> List[TimeSlot]:
        all_slots = self.read_all_slots()
        return [slot for slot in all_slots if slot.is_free()]
    
    def get_slots_by_date(self, slot_date: date) -> List[TimeSlot]:
        all_slots = self.read_all_slots()
        return [slot for slot in all_slots if slot.slot_date == slot_date]
    
    def close(self) -> None:
        if self.workbook:
            self.workbook.close()
            logger.info("Файл расписания закрыт")
